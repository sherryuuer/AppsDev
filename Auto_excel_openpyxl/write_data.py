import openpyxl as xl


WS_WORKSHEET = "作業シート"
WS_MANAGERSHEET = "管理シート"
WS_DELETE = "削除済み"

# フォーマットExcelの読み込み
wb = xl.load_workbook('ToolBox\Auto_excel_openpyxl\ユーザ管理.xlsx')

def write_data(target_sheet=WS_DELETE):
    ws_work = wb[WS_WORKSHEET]
    target_sheet = wb[target_sheet]

    # 作業シートからデータを取得
    data_to_copy = []
    for row in ws_work.iter_rows(min_row=3, values_only=True):
        data_to_copy.append(row)

    # 管理シートのB2セルにデータを設定
    for data_row in data_to_copy:
        # B列にデータを設定
        for i, cell_value in enumerate(data_row, start=1):
            target_sheet.cell(row=2, column=i, value=cell_value)  # B2セルにデータを設定


    # ワークブックを保存
    wb.save('ToolBox\Auto_excel_openpyxl\ユーザ管理.xlsx')

# 関数を呼び出してデータをコピー
write_data()



